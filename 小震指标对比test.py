import openpyxl

# 创建workbook对象
wb = openpyxl.Workbook()
# # 删除默认创建的一个sheet页
ws = wb['Sheet']
wb.remove(ws)
# # 给sheet页命名
sheetName = "YJK Etabs 对比"
# 创建sheet页
ws = wb.create_sheet(sheetName)
# 合并单元格
ws.merge_cells('A1:B1')
# 写入方式1：（行、列、值）
ws.cell(row=1, column=1, value="计算软件")
ws.cell(row=1, column=3).value = "规范限值"


filePath = 'YJK Etabs 对比.xlsx'
wb.save(filePath)
